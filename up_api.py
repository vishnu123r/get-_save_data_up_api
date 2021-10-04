import requests
from pprint import pprint
import pickle 
from openpyxl import load_workbook, workbook
import pandas as pd
import sys
import calendar
import datetime
from dateutil import relativedelta 
from dotenv import load_dotenv
import os

load_dotenv()
up_api_key = os.getenv('up_api_key')
current_account = os.getenv('current')

def get_query_time_interval():
    ### This function returns the time period of interest based on the current date
    current_date = datetime.datetime.now()
    one_month = relativedelta.relativedelta(months = 1)
    query_date = current_date - one_month
    query_month = query_date.month
    query_year = query_date.year
    query_days = calendar.monthrange(query_year, query_month)[1]

    since = '{}-{}-01T00:00:00Z'.format(query_year, str(query_month).zfill(2))
    until = '{}-{}-{}T00:00:00Z'.format(query_year, str(query_month).zfill(2), query_days)

    return since, until, query_month, query_year, query_days

def get_up_data(since, until):
    try:
        url = "https://api.up.com.au/api/v1"
        payload = {'filter[since]': since, 'filter[until]': until, 'page[size]':'100'}
        header = {'Authorization': up_api_key}
        print(header)
        r = requests.get(url+"/transactions", headers = header, params = payload)
        if r.ok:
            transaction_list = [r.json()]
            x = 0
            if r.json()['links']['next'] is None:
                x = 1
            while x == 0:
                r = requests.get(r.json()['links']['next'], headers = header, params = payload)
                if r.json()['links']['next'] is None:
                    x = 1
                transaction_list.append(r.json())

            return transaction_list
        else: 
            print("Bad response from Up Api = ", r.json())
            sys.exit()

    except Exception as e:
            print("error: {error_message}". format(error_message=str(e)))
 
def sum_transactions(output):
    ### This function is used to sum up all the transaction values from the response dictionary
    category_dict = {}
    for response in output:
        for transaction in response['data']:
            id = transaction['relationships']['account']['data']['id']
            if id == current_account:
                category = transaction['relationships']['category']['data']
                if category is not None:
                    category = category['id']
                    if category not in category_dict:
                        category_dict[category] = round(float(transaction['attributes']['amount']['value']), 2)
                    else:
                        category_dict[category] += round(float(transaction['attributes']['amount']['value']), 2)

    for key, value in category_dict.items():
        category_dict[key] = round(value,2)

    return category_dict

def check_new_types(category_dict):
    ### Used to check if there are new categories 
    types_considered = [  'groceries', 'health-and-medical', 'life-admin', 'restaurants-and-cafes', 'rent-and-mortgage', 'mobile-phone', 'takeaway', 'holidays-and-travel', 'hair-and-beauty', 'education-and-student-loans', 'news-magazines-and-books', 'public-transport', 'homeware-and-appliances', 'games-and-software', 'taxis-and-share-cars', 'clothing-and-accessories', 'gifts-and-charity', 'home-maintenance-and-improvements']
    diff_set = str(set(list(category_dict.keys())) - set(types_considered))
    error_message = 'There are new names in the incoming file ' + diff_set
    if diff_set != 'set()':
        sys.exit(error_message)

def category_selector(x):
    ### To arrange categories. Used with the function - order_df
    good_life_except = ('games-and-software', 'holidays-and-travel', 'restaurants-and-cafes', )
    take_away = ('takeaway')
    home_except = ('homeware-and-appliances', 'home-maintenance-and-improvements')
    personal = ('health-and-medical', 'life-admin', 'mobile-phone', 'hair-and-beauty', 'education-and-student-loans', 'news-magazines-and-books', 'clothing-and-accessories', 'gifts-and-charity')
    transport = ('public-transport', 'taxis-and-share-cars')

    if  x in  good_life_except:
        return 'Good life (except Takeaway)'
    elif x in take_away:
        return 'Takeaway'
    elif x in home_except:
        return 'Home (Other)'
    elif x == 'groceries':
        return 'Groceries'
    elif x == 'rent-and-mortgage':
        return 'Rent'
    elif x in personal:
        return 'Personal'
    else:
        return 'Transport'

def order_df(category_dict, query_month, query_year, query_days):
    ### This function converts the dictionary to dataframe (to be used to write excel)
    df = pd.DataFrame.from_dict([category_dict])
    df = df.transpose().reset_index()
    df.columns = ['type', 'amount']
    df['category'] = df['type'].apply(lambda x: category_selector(x))
    df = df.groupby(by='category').sum()
    df['amount'] = round(df['amount'] *-1,0)
    #Add zero value if there are no values in the category
    for cat in ['Good life (except Takeaway)', 'Groceries', 'Personal', 'Rent', 'Takeaway', 'Transport', 'Home (Other)']:
        if cat not in df.index:
            df.loc[cat] = 0
    
    df = df.transpose()
    df['total'] = df.sum(axis=1)
    df['month'] = query_month
    df['year'] = query_year
    df['days'] = query_days
    df['income'] = None
    df['hourly_rate'] = None 
    df['hours_per_week'] = df['income']/df['hourly_rate']
    df['good_life_%'] = round(df['Good life (except Takeaway)']*100/df['total'],0)
    df['personal_%'] = round(df['Personal']*100/df['total'],0)
    df['rent_%'] = round(df['Rent']*100/df['total'],0)
    df['food_%'] = round((df['Groceries'] + df['Takeaway'])*100/df['total'],0)
    df['home_%'] = round(df['Transport']*100/df['total'],0)
    df['Transport_%'] = round(df['Transport']*100/df['total'],0)
    column_order = ['year','month', 'days', 'income', 'hourly_rate', 'hours_per_week', 'total', 'Good life (except Takeaway)', 'Takeaway', 'Personal', 'Rent', 'Groceries', 'Home (Other)', 
    'Transport', 'good_life_%', 'personal_%', 'rent_%', 'food_%', 'home_%', 'Transport_%']
    df = df.reindex(columns = column_order)

    return df

def convert_df_excel(df, file_name = 'demo.xlsx', sheet_name = 'Money'):
    ### This function converts df to excel file
    book = load_workbook(file_name) #Open_workbook in pyxl
    try:
        writer = pd.ExcelWriter(file_name, engine='openpyxl') #Open in Excel writer
        writer.book = book #connect both pyxl and pandas (excel_writer)
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        startrow = book[sheet_name].max_row
        start_column = 1

        #Check if the data already exists
        last_month = book[sheet_name].cell(row = startrow, column=3).value
        last_year = book[sheet_name].cell(row = startrow, column=2).value
        if last_month == query_month and last_year == query_year:
            book.save(file_name)
            error_string = "The data already exists for month {} and year {}".format(query_month, query_year)
            print(error_string)
            df.to_excel(writer, sheet_name, startrow=startrow-1, startcol = start_column, header = False, index = False)
            sys.exit('Data has been replaced')    

        df.to_excel(writer, sheet_name, startrow=startrow, startcol = start_column, header = False, index = False)
        
        print("Data has been inserted")
        book.save(file_name)

    except Exception as e:
        print("error: {error_message}". format(error_message=str(e)))
        book.save(file_name)

if __name__ == '__main__':
    since, until, query_month, query_year, query_days = get_query_time_interval()
    output = get_up_data(since, until)
    category_dict = sum_transactions(output)
    check_new_types(category_dict)
    df = order_df(category_dict, query_month, query_year, query_days)
    convert_df_excel(df)